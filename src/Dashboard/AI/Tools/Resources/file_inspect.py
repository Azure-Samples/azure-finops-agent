"""
file_inspect.py — uploaded-file preview + query helper for the Azure FinOps Agent.

Invocation:
    python3 file_inspect.py
    stdin: JSON request { mode, path, kind, ... }
    stdout: JSON response { ok, ... }   (single line, always JSON)
    stderr: human-readable error (only on hard failure)

Modes:
    preview        Initial inspection — schema + first N rows / chars (used at upload time)
    schema         Re-emit just the schema (cheap)
    head           First N rows / chars
    tail           Last N rows / chars
    slice          Rows offset..offset+count (or text bytes)
    text_range     For txt/pdf — substring(start, length)
    count          Row count (and column count where applicable)
    filter         CSV/JSON-records: rows where col {op} value (op in eq, ne, gt, lt, ge, le, contains)
    aggregate      CSV/JSON-records: group_by + agg(sum|mean|min|max|count) on a numeric column
    json_path      JSON: navigate dot/bracket path, return that subtree (truncated)

All responses cap payload size (rows, characters) so the model never sees more
than a small chunk per call. The model is expected to make multiple calls.
"""
from __future__ import annotations
import io
import json
import math
import os
import sys
import traceback
import warnings
from collections import deque
from typing import Any

# stdout MUST be pure JSON — route every warning to stderr so a stray
# pandas/pyarrow DeprecationWarning can't corrupt the response.
warnings.simplefilter("ignore")
warnings.showwarning = lambda *a, **kw: None

# Hard caps — these protect the LLM context window
MAX_ROWS_PER_CALL = 200
MAX_TEXT_CHARS_PER_CALL = 8000
PREVIEW_ROWS = 50
PREVIEW_TEXT_CHARS = 5000
PREVIEW_JSON_ITEMS = 10
SCHEMA_MAX_KEYS = 200
SCHEMA_MAX_DEPTH = 6


def _ok(**payload: Any) -> dict:
    return {"ok": True, **payload}


def _err(message: str, **extra: Any) -> dict:
    return {"ok": False, "error": message, **extra}


def _clean(value: Any) -> Any:
    """Recursively replace NaN/Inf (invalid JSON) with None and downcast numpy scalars."""
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, dict):
        return {k: _clean(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_clean(v) for v in value]
    # numpy / pandas scalars
    try:
        import numpy as _np  # local import to avoid hard dep at module load
        if isinstance(value, _np.generic):
            v = value.item()
            return _clean(v)
    except Exception:
        pass
    return value


# ------------------------------------------------------------------ JSON utils

def _json_schema(value: Any, depth: int = 0) -> Any:
    if depth >= SCHEMA_MAX_DEPTH:
        return "..."
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, str):
        return "string"
    if isinstance(value, list):
        if not value:
            return ["empty"]
        # represent as one-element list of the merged item shape
        sample = value[0]
        return [_json_schema(sample, depth + 1)]
    if isinstance(value, dict):
        out = {}
        for i, (k, v) in enumerate(value.items()):
            if i >= SCHEMA_MAX_KEYS:
                out["..."] = f"+{len(value) - SCHEMA_MAX_KEYS} more keys"
                break
            out[k] = _json_schema(v, depth + 1)
        return out
    return type(value).__name__


def _json_path_get(root: Any, path: str) -> Any:
    """Very small dot/bracket navigator: a.b[0].c"""
    if not path:
        return root
    cur = root
    token = ""
    i = 0
    parts: list[str | int] = []
    while i < len(path):
        ch = path[i]
        if ch == ".":
            if token:
                parts.append(token)
                token = ""
            i += 1
        elif ch == "[":
            if token:
                parts.append(token)
                token = ""
            j = path.find("]", i)
            if j < 0:
                raise ValueError(f"unterminated [ at {i}")
            parts.append(int(path[i + 1 : j]))
            i = j + 1
        else:
            token += ch
            i += 1
    if token:
        parts.append(token)
    for p in parts:
        if isinstance(p, int):
            cur = cur[p]
        else:
            cur = cur[p]
    return cur


# --------------------------------------------------------------------- handlers

def _handle_text(req: dict, raw: bytes) -> dict:
    text = raw.decode("utf-8", errors="replace")
    total = len(text)
    mode = req["mode"]
    if mode in ("preview", "head"):
        n = min(int(req.get("chars", PREVIEW_TEXT_CHARS)), MAX_TEXT_CHARS_PER_CALL)
        return _ok(kind="txt", total_chars=total, chunk=text[:n], preview_chars=n)
    if mode == "tail":
        n = min(int(req.get("chars", PREVIEW_TEXT_CHARS)), MAX_TEXT_CHARS_PER_CALL)
        return _ok(kind="txt", total_chars=total, chunk=text[-n:], preview_chars=n)
    if mode == "text_range":
        start = max(0, int(req.get("start", 0)))
        length = min(int(req.get("length", PREVIEW_TEXT_CHARS)), MAX_TEXT_CHARS_PER_CALL)
        return _ok(kind="txt", total_chars=total, chunk=text[start : start + length])
    if mode == "count":
        return _ok(kind="txt", total_chars=total, total_lines=text.count("\n") + 1)
    if mode == "schema":
        return _ok(kind="txt", total_chars=total)
    return _err(f"mode '{mode}' not supported for txt")


def _handle_json(req: dict, raw: bytes) -> dict:
    try:
        data = json.loads(raw.decode("utf-8", errors="replace"))
    except json.JSONDecodeError as e:
        return _err(f"invalid JSON: {e}")
    mode = req["mode"]

    if mode in ("preview", "schema"):
        if isinstance(data, list):
            return _ok(
                kind="json",
                shape="array",
                length=len(data),
                schema=_json_schema(data[:1]),
                first_items=data[: PREVIEW_JSON_ITEMS],
            )
        return _ok(
            kind="json",
            shape="object" if isinstance(data, dict) else type(data).__name__,
            schema=_json_schema(data),
            sample=data if not isinstance(data, (list, dict)) else None,
        )

    if mode == "json_path":
        try:
            sub = _json_path_get(data, req.get("path", ""))
        except Exception as e:
            return _err(f"json_path error: {e}")
        # truncate
        if isinstance(sub, list):
            return _ok(kind="json", path=req.get("path", ""), length=len(sub), items=sub[:MAX_ROWS_PER_CALL])
        return _ok(kind="json", path=req.get("path", ""), value=sub if not isinstance(sub, dict) else dict(list(sub.items())[:SCHEMA_MAX_KEYS]))

    if mode in ("head", "tail", "slice"):
        if not isinstance(data, list):
            return _err("head/tail/slice require a JSON array root")
        offset = int(req.get("offset", 0))
        count = min(int(req.get("count", 50)), MAX_ROWS_PER_CALL)
        if mode == "head":
            chunk = data[:count]
        elif mode == "tail":
            chunk = data[-count:]
        else:
            chunk = data[offset : offset + count]
        return _ok(kind="json", length=len(data), offset=offset if mode == "slice" else 0, items=chunk)

    if mode == "count":
        if isinstance(data, list):
            return _ok(kind="json", length=len(data))
        if isinstance(data, dict):
            return _ok(kind="json", keys=len(data))
        return _ok(kind="json", scalar=True)

    if mode in ("filter", "aggregate"):
        # Treat list-of-objects as a tabular dataset and reuse pandas
        if not isinstance(data, list) or not data or not isinstance(data[0], dict):
            return _err("filter/aggregate require a JSON array of objects")
        import pandas as pd
        df = pd.DataFrame(data)
        return _df_query(df, req, kind="json")

    return _err(f"mode '{mode}' not supported for json")


def _df_query(df, req: dict, kind: str) -> dict:
    import pandas as pd  # noqa: F401
    mode = req["mode"]

    if mode == "filter":
        col = req["column"]
        op = req.get("op", "eq")
        val = req.get("value")
        limit = min(int(req.get("limit", 50)), MAX_ROWS_PER_CALL)
        if col not in df.columns:
            return _err(f"unknown column '{col}'", columns=list(df.columns))
        s = df[col]
        try:
            if op == "eq":
                mask = s == val
            elif op == "ne":
                mask = s != val
            elif op == "gt":
                mask = pd.to_numeric(s, errors="coerce") > float(val)
            elif op == "lt":
                mask = pd.to_numeric(s, errors="coerce") < float(val)
            elif op == "ge":
                mask = pd.to_numeric(s, errors="coerce") >= float(val)
            elif op == "le":
                mask = pd.to_numeric(s, errors="coerce") <= float(val)
            elif op == "contains":
                mask = s.astype(str).str.contains(str(val), case=False, na=False)
            else:
                return _err(f"unknown op '{op}'")
        except Exception as e:
            return _err(f"filter failed: {e}")
        sub = df[mask].head(limit)
        return _ok(kind=kind, total_matches=int(mask.sum()), rows=sub.to_dict(orient="records"))

    if mode == "aggregate":
        gb = req.get("group_by")
        agg = req.get("agg", "sum")
        col = req.get("column")
        limit = min(int(req.get("limit", 50)), MAX_ROWS_PER_CALL)
        if col not in df.columns:
            return _err(f"unknown column '{col}'", columns=list(df.columns))
        series = pd.to_numeric(df[col], errors="coerce")
        if gb:
            if gb not in df.columns:
                return _err(f"unknown group_by column '{gb}'", columns=list(df.columns))
            grouped = series.groupby(df[gb])
            result = getattr(grouped, agg)()
            result = result.sort_values(ascending=False).head(limit)
            return _ok(kind=kind, agg=agg, group_by=gb, column=col, rows=[{gb: k, agg: float(v) if pd.notna(v) else None} for k, v in result.items()])
        scalar = getattr(series, agg)()
        return _ok(kind=kind, agg=agg, column=col, value=float(scalar) if pd.notna(scalar) else None)

    return _err(f"mode '{mode}' not supported here")


def _handle_csv(req: dict, raw: bytes) -> dict:
    import pandas as pd
    sep = req.get("sep", ",")
    try:
        df = pd.read_csv(io.BytesIO(raw), sep=sep, low_memory=False)
    except Exception as e:
        try:
            df = pd.read_csv(io.BytesIO(raw), sep=None, engine="python")
        except Exception as e2:
            return _err(f"csv parse failed: {e}; fallback: {e2}")
    return _tabular_response(df, req, kind="csv")


def _handle_xlsx(req: dict, path: str) -> dict:
    # Importing pandas costs 10-15 seconds on the Windows ARM64 development
    # path and was repeated for every workbook query. openpyxl is already the
    # XLSX engine and can stream these operations directly without another
    # heavyweight import or materialising entire workbooks as data frames.
    from openpyxl import load_workbook

    def headers_for(values) -> list[str]:
        headers: list[str] = []
        seen: dict[str, int] = {}
        for index, value in enumerate(values):
            base = str(value).strip() if value is not None and str(value).strip() else f"Column{index + 1}"
            seen[base] = seen.get(base, 0) + 1
            headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
        return headers

    def rows_for(ws):
        iterator = ws.iter_rows(values_only=True)
        headers = headers_for(next(iterator, ()))
        for values in iterator:
            values = tuple(values[: len(headers)])
            if not any(value is not None for value in values):
                continue
            if len(values) < len(headers):
                values += (None,) * (len(headers) - len(values))
            yield headers, values

    def row_dict(headers: list[str], values: tuple) -> dict:
        return {headers[index]: values[index] for index in range(len(headers))}

    def number(value: Any) -> float | None:
        if isinstance(value, bool) or value is None:
            return None
        if isinstance(value, (int, float)):
            parsed = float(value)
        else:
            try:
                parsed = float(str(value).replace(",", "").strip())
            except (TypeError, ValueError):
                return None
        return parsed if math.isfinite(parsed) else None

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        if req["mode"] == "workbook":
            sheets = []
            for ws in wb.worksheets:
                iterator = ws.iter_rows(values_only=True)
                headers = headers_for(next(iterator, ()))
                stats = [None] * len(headers)
                total_rows = 0
                for values in iterator:
                    values = tuple(values[: len(headers)])
                    if not any(value is not None for value in values):
                        continue
                    total_rows += 1
                    for index, value in enumerate(values):
                        parsed = number(value)
                        if parsed is None:
                            continue
                        current = stats[index]
                        if current is None:
                            stats[index] = {"count": 1, "sum": parsed, "min": parsed, "max": parsed}
                        else:
                            current["count"] += 1
                            current["sum"] += parsed
                            current["min"] = min(current["min"], parsed)
                            current["max"] = max(current["max"], parsed)

                numeric_summaries = {}
                for index, current in enumerate(stats):
                    if current is None or len(numeric_summaries) >= 20:
                        continue
                    numeric_summaries[headers[index]] = {
                        "count": current["count"],
                        "sum": current["sum"],
                        "min": current["min"],
                        "max": current["max"],
                        "mean": current["sum"] / current["count"],
                    }
                sheets.append({
                    "name": ws.title,
                    "total_rows": total_rows,
                    "total_columns": len(headers),
                    "columns": headers[:SCHEMA_MAX_KEYS],
                    "numeric_summaries": numeric_summaries,
                })
            return _ok(
                kind="xlsx",
                sheet_count=len(sheets),
                total_rows=sum(sheet["total_rows"] for sheet in sheets),
                sheets=sheets,
            )

        sheet = req.get("sheet") or sheet_names[0]
        if sheet not in sheet_names:
            return _err(f"unknown sheet '{sheet}'", sheets=sheet_names)
        ws = wb[sheet]
        mode = req["mode"]
        rows = list(rows_for(ws))
        headers = rows[0][0] if rows else headers_for(next(ws.iter_rows(values_only=True), ()))
        data = [values for _, values in rows]

        if mode in ("preview", "schema"):
            sample = data[: min(int(req.get("rows", PREVIEW_ROWS)), MAX_ROWS_PER_CALL)]
            dtypes = {}
            for index, header in enumerate(headers):
                types = {type(row[index]).__name__ for row in sample if row[index] is not None}
                dtypes[header] = next(iter(types)) if len(types) == 1 else "object"
            payload = _ok(
                kind="xlsx",
                total_rows=len(data),
                total_columns=len(headers),
                columns=headers,
                dtypes=dtypes,
            )
            if mode == "preview":
                payload["rows"] = [row_dict(headers, row) for row in sample]
                payload["preview_rows"] = len(sample)
        elif mode == "count":
            payload = _ok(kind="xlsx", total_rows=len(data), total_columns=len(headers))
        elif mode in ("head", "tail", "slice"):
            count = min(int(req.get("count", PREVIEW_ROWS)), MAX_ROWS_PER_CALL)
            if mode == "head":
                selected = data[:count]
                offset = 0
            elif mode == "tail":
                selected = list(deque(data, maxlen=count))
                offset = max(0, len(data) - len(selected))
            else:
                offset = max(0, int(req.get("offset", 0)))
                selected = data[offset : offset + count]
            payload = _ok(kind="xlsx", offset=offset, rows=[row_dict(headers, row) for row in selected])
        elif mode == "aggregate":
            column = req.get("column")
            group_by = req.get("group_by")
            agg = req.get("agg", "sum")
            if column not in headers:
                return _err(f"unknown column '{column}'", columns=headers)
            if group_by and group_by not in headers:
                return _err(f"unknown group_by column '{group_by}'", columns=headers)
            if agg not in ("sum", "mean", "min", "max", "count"):
                return _err(f"unknown aggregate '{agg}'")
            column_index = headers.index(column)
            group_index = headers.index(group_by) if group_by else None
            groups: dict[Any, list[float]] = {}
            for row in data:
                parsed = number(row[column_index])
                if parsed is None:
                    continue
                key = row[group_index] if group_index is not None else None
                if group_index is not None and key is None:
                    continue
                groups.setdefault(key, []).append(parsed)

            def aggregate(values: list[float]) -> float | int | None:
                if not values:
                    return 0 if agg in ("sum", "count") else None
                if agg == "sum":
                    return sum(values)
                if agg == "mean":
                    return sum(values) / len(values)
                if agg == "min":
                    return min(values)
                if agg == "max":
                    return max(values)
                return len(values)

            if group_by:
                limit = min(int(req.get("limit", 50)), MAX_ROWS_PER_CALL)
                values = [(key, aggregate(group)) for key, group in groups.items()]
                values.sort(key=lambda item: item[1] if item[1] is not None else float("-inf"), reverse=True)
                payload = _ok(kind="xlsx", agg=agg, group_by=group_by, column=column,
                              rows=[{group_by: key, agg: value} for key, value in values[:limit]])
            else:
                payload = _ok(kind="xlsx", agg=agg, column=column, value=aggregate(groups.get(None, [])))
        elif mode == "filter":
            column = req.get("column")
            op = req.get("op", "eq")
            expected = req.get("value")
            if column not in headers:
                return _err(f"unknown column '{column}'", columns=headers)
            column_index = headers.index(column)
            limit = min(int(req.get("limit", 50)), MAX_ROWS_PER_CALL)
            matches = []
            total_matches = 0
            for row in data:
                actual = row[column_index]
                try:
                    if op == "contains":
                        matched = str(expected).lower() in str(actual).lower()
                    elif op in ("gt", "lt", "ge", "le"):
                        left, right = number(actual), number(expected)
                        matched = left is not None and right is not None and {
                            "gt": left > right,
                            "lt": left < right,
                            "ge": left >= right,
                            "le": left <= right,
                        }[op]
                    elif op == "eq":
                        matched = actual == expected or str(actual).lower() == str(expected).lower()
                    elif op == "ne":
                        matched = not (actual == expected or str(actual).lower() == str(expected).lower())
                    else:
                        return _err(f"unknown op '{op}'")
                except Exception:
                    matched = False
                if matched:
                    total_matches += 1
                    if len(matches) < limit:
                        matches.append(row_dict(headers, row))
            payload = _ok(kind="xlsx", total_matches=total_matches, rows=matches)
        else:
            return _err(f"mode '{mode}' not supported for xlsx")

        payload["sheet"] = sheet
        payload["sheets"] = sheet_names
        return payload
    finally:
        wb.close()


def _handle_parquet(req: dict, path: str) -> dict:
    import pandas as pd
    df = pd.read_parquet(path)
    return _tabular_response(df, req, kind="parquet")


def _tabular_response(df, req: dict, kind: str) -> dict:
    mode = req["mode"]
    cols = list(df.columns)
    dtypes = {c: str(df[c].dtype) for c in cols}

    if mode in ("preview",):
        head_n = min(int(req.get("rows", PREVIEW_ROWS)), MAX_ROWS_PER_CALL)
        return _ok(
            kind=kind,
            total_rows=int(len(df)),
            total_columns=len(cols),
            columns=cols,
            dtypes=dtypes,
            rows=df.head(head_n).to_dict(orient="records"),
            preview_rows=head_n,
        )
    if mode == "schema":
        return _ok(kind=kind, total_rows=int(len(df)), total_columns=len(cols), columns=cols, dtypes=dtypes)
    if mode == "count":
        return _ok(kind=kind, total_rows=int(len(df)), total_columns=len(cols))
    if mode == "head":
        n = min(int(req.get("count", PREVIEW_ROWS)), MAX_ROWS_PER_CALL)
        return _ok(kind=kind, rows=df.head(n).to_dict(orient="records"))
    if mode == "tail":
        n = min(int(req.get("count", PREVIEW_ROWS)), MAX_ROWS_PER_CALL)
        return _ok(kind=kind, rows=df.tail(n).to_dict(orient="records"))
    if mode == "slice":
        offset = max(0, int(req.get("offset", 0)))
        count = min(int(req.get("count", PREVIEW_ROWS)), MAX_ROWS_PER_CALL)
        return _ok(kind=kind, offset=offset, rows=df.iloc[offset : offset + count].to_dict(orient="records"))
    if mode in ("filter", "aggregate"):
        return _df_query(df, req, kind=kind)
    return _err(f"mode '{mode}' not supported for {kind}")


def _handle_pdf(req: dict, path: str) -> dict:
    try:
        from pdfminer.high_level import extract_text
    except Exception as e:
        return _err(f"pdfminer not available: {e}")
    text = extract_text(path) or ""
    # piggyback on text handler semantics
    raw = text.encode("utf-8")
    return _handle_text(req, raw) | {"kind": "pdf"}


# ------------------------------------------------------------------------ main

def main() -> int:
    try:
        req = json.loads(sys.stdin.read() or "{}")
    except json.JSONDecodeError as e:
        print(json.dumps(_err(f"bad request json: {e}")))
        return 0

    path = req.get("path")
    kind = (req.get("kind") or "").lower()
    if not path or not os.path.exists(path):
        print(json.dumps(_err("file not found", path=path)))
        return 0

    try:
        if kind in ("xlsx", "xls"):
            resp = _handle_xlsx(req, path)
        elif kind == "parquet":
            resp = _handle_parquet(req, path)
        elif kind == "pdf":
            resp = _handle_pdf(req, path)
        else:
            with open(path, "rb") as f:
                raw = f.read()
            if kind == "csv" or kind == "tsv":
                if kind == "tsv":
                    req.setdefault("sep", "\t")
                resp = _handle_csv(req, raw)
            elif kind == "json":
                resp = _handle_json(req, raw)
            else:  # txt and unknown → text fallback
                resp = _handle_text(req, raw)
    except Exception as e:
        resp = _err(f"{type(e).__name__}: {e}", trace=traceback.format_exc().splitlines()[-5:])

    # Hard cap stdout size — keep the LLM context small
    try:
        out = json.dumps(_clean(resp), default=str, allow_nan=False)
    except (ValueError, TypeError) as e:
        out = json.dumps({"ok": False, "error": f"json serialize failed: {e}"})
    if len(out) > 64_000:
        out = json.dumps({"ok": False, "error": "response too large; narrow your query (use head/slice with smaller count, or filter/aggregate)", "size": len(out)})
    print(out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
